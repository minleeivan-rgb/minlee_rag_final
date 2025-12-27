#!/usr/bin/env python3
"""
從歷史檔案中提取文字 (與 utils.py 完美同步版)
支援：Excel、文字型 PDF、掃描型 PDF (OCR)
"""

import os
import json
from openpyxl import load_workbook
from tqdm import tqdm
import pdfplumber
import uuid
# 核心修正：這裡的名字必須與 utils.py 裡的 AzureOpenAIAPI 完全一致
from utils import GeminiAPI, AzureOpenAIAPI, load_config, print_progress

# OCR 相關套件（延遲載入，避免沒用到時報錯）
ocr_reader = None

def get_ocr_reader():
    """延遲載入 EasyOCR（第一次使用時才載入，節省啟動時間）"""
    global ocr_reader
    if ocr_reader is None:
        print("[INFO] 首次載入 OCR 模型，請稍候...")
        import easyocr
        ocr_reader = easyocr.Reader(['ch_tra', 'en'], gpu=False)  # 繁體中文 + 英文
        print("[INFO] OCR 模型載入完成")
    return ocr_reader

def extract_bom_text_from_excel(excel_path):
    """提取 Excel 中的文字（支援 Numbers 轉換的檔案）"""
    try:
        wb = load_workbook(excel_path, data_only=True)
        all_text = ""
        bom_items = []
        
        # 遍歷所有工作表（Numbers 轉換可能產生多個工作表）
        for sheet in wb.worksheets:
            sheet_text = ""
            
            for row in range(1, sheet.max_row + 1):
                row_vals = []
                for col in range(1, sheet.max_column + 1):
                    val = sheet.cell(row, col).value
                    if val:
                        row_vals.append(str(val).strip())
                
                line_text = " ".join(row_vals)
                if line_text:
                    sheet_text += line_text + " "
                    # 如果有 # 字號，判定為零件項
                    if "#" in line_text or "＃" in line_text:
                        bom_items.append({
                            "number": row_vals[0] if row_vals else "", 
                            "full_text": line_text
                        })
            
            # 跳過 Numbers 轉換說明的工作表
            if "Numbers" in sheet_text and "輸出" in sheet_text:
                continue
            if "此文件從" in sheet_text:
                continue
                
            all_text += sheet_text
        
        return {
            'filename': os.path.basename(excel_path),
            'bom_items': bom_items,
            'full_text': all_text.strip()
        }
    except Exception as e:
        print(f"❌ Excel Error ({os.path.basename(excel_path)}): {e}")
        return None

def extract_bom_text_from_pdf(pdf_path):
    """提取 PDF 中的文字（支援掃描型 PDF）"""
    try:
        all_text = ""
        
        # 第一步：嘗試用 pdfplumber 提取文字（適用於文字型 PDF）
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text: 
                    all_text += text + "\n"
        
        # 第二步：如果提取的文字太少，可能是掃描型 PDF，改用 OCR
        if len(all_text.strip()) < 50:
            print(f"⚠️  [{os.path.basename(pdf_path)}] 文字太少，嘗試 OCR...")
            all_text = extract_text_with_ocr(pdf_path)
        
        return {
            'filename': os.path.basename(pdf_path),
            'bom_items': [], 
            'full_text': all_text.strip()
        }
    except Exception as e:
        print(f"❌ PDF Error ({os.path.basename(pdf_path)}): {e}")
        return None


def extract_text_with_ocr(pdf_path):
    """使用 OCR 從掃描型 PDF 提取文字"""
    try:
        import fitz  # pymupdf
        
        all_text = ""
        reader = get_ocr_reader()
        
        # 打開 PDF
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        
        print(f"[INFO] 開始 OCR 處理 {total_pages} 頁...")
        
        for page_num in range(total_pages):
            # 把 PDF 頁面轉成圖片
            page = doc.load_page(page_num)
            # 提高解析度以獲得更好的 OCR 效果
            mat = fitz.Matrix(2.0, 2.0)  # 放大 2 倍
            pix = page.get_pixmap(matrix=mat)
            
            # 轉成 PIL Image
            img_data = pix.tobytes("png")
            
            # 使用 EasyOCR 辨識
            results = reader.readtext(img_data)
            
            # 提取文字
            page_text = " ".join([result[1] for result in results])
            all_text += page_text + "\n"
            
            print(f"  ✓ 第 {page_num + 1}/{total_pages} 頁 OCR 完成")
        
        doc.close()
        print(f"✅ OCR 完成，提取 {len(all_text)} 字元")
        
        return all_text
        
    except ImportError:
        print("❌ 缺少 pymupdf 套件，請執行：pip install pymupdf")
        return ""
    except Exception as e:
        print(f"❌ OCR Error: {e}")
        return ""

def process_all_files(history_folder, output_json='extracted_data.json'):
    """批量處理檔案並生成向量存入 JSON"""
    config = load_config()
    # 這裡會呼叫 utils.py 裡的類別
    azure_ai = AzureOpenAIAPI()
    gemini = GeminiAPI()
    
    all_files = [os.path.join(history_folder, f) for f in os.listdir(history_folder) 
                 if f.endswith(('.xlsx', '.xls', '.pdf'))]
    
    if not all_files:
        print(f"❌ 資料夾內找不到任何支援的檔案: {history_folder}")
        return []

    all_data = []
    for file_path in tqdm(all_files, desc="Processing files"):
        if file_path.endswith('.pdf'):
            data = extract_bom_text_from_pdf(file_path)
        else:
            data = extract_bom_text_from_excel(file_path)
            
        if data and data['full_text']:
            # 1. 提取型號（用於顯示和輔助搜尋）
            model_hint = gemini.enhance_bom_text(data['full_text'])
            
            # 2. 用完整 BOM 內容生成向量（重要！查詢時也用 full_text）
            vector = azure_ai.get_embedding(data['full_text'])
            
            if vector:
                doc = {
                    'document_id': str(uuid.uuid4()),
                    'filename': data['filename'],
                    'bom_items': data['bom_items'],
                    'full_text': data['full_text'],
                    'model_hint': model_hint,  # 額外儲存型號提示
                    'vector': vector,
                    'is_primary': True
                }
                all_data.append(doc)
                print(f"  ✓ {data['filename']} (型號: {model_hint})")
            
    # 儲存結果
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    
    print(f"✅ 成功處理 {len(all_data)} 筆資料並儲存至 {output_json}")
    return all_data

if __name__ == '__main__':
    config = load_config()
    history_folder = config['PATHS']['history_excel_folder']
    process_all_files(history_folder)