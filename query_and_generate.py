import os
import sys
from pymongo import MongoClient
from utils import AzureOpenAIAPI, GeminiAPI, load_config, print_progress
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def query_similar_boms(query_vector, config, top_k=3):
    """用向量相似度搜尋最相似的 BOM 模板"""
    try:
        client = MongoClient(config['MONGODB']['connection_string'])
        db = client[config['MONGODB']['database_name']]
        collection = db[config['MONGODB']['collection_name']]
        
        print_progress("正在用向量搜尋相似模板...")
        
        # 向量搜尋
        results = list(collection.aggregate([
            {
                "$vectorSearch": {
                    "index": config['MONGODB']['vector_index_name'],
                    "path": "vector",
                    "queryVector": query_vector,
                    "numCandidates": 100,
                    "limit": top_k
                }
            },
            {
                "$project": {
                    "_id": 0, "filename": 1, "bom_items": 1, "full_text": 1,
                    "model_hint": 1,
                    "score": {"$meta": "vectorSearchScore"}
                }
            }
        ]))
        client.close()
        
        # 印出搜尋結果
        print("\n[向量搜尋結果 - 按相似度排序]:")
        for i, r in enumerate(results):
            model = r.get('model_hint', '未知')
            print(f"  {i+1}. {r['filename']} (相似度: {r['score']:.4f}, 型號: {model})")
        print()
        
        return results
    except Exception as e:
        print(f"❌ 資料庫搜尋發生錯誤: {e}")
        return []

def create_styled_excel(steps, product_name, output_path):
    """建立帶有專業格式與大照片框的 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "SOP"
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    ws.merge_cells('A1:E1')
    ws['A1'] = f"產品組裝指導書 - {product_name}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ["步驟", "標題", "照片示意圖 (預留位)", "組裝詳細說明", "注意事項"]
    ws.append(headers)
    
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row_idx = 3
    for s in steps:
        ws.cell(row=row_idx, column=1, value=s.get('step_number')).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_idx, column=2, value=s.get('title')).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_idx, column=3, value="[ 預留照片位置 ]").alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row_idx, column=4, value=s.get('description')).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.cell(row=row_idx, column=5, value=s.get('notes')).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        for col in range(1, 6): ws.cell(row=row_idx, column=col).border = border
        ws.row_dimensions[row_idx].height = 180 # 超大照片格
        row_idx += 1

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 45 # 寬照片格
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 30
    wb.save(output_path)

def main():
    if len(sys.argv) < 2:
        print("用法: python3 query_and_generate.py <BOM檔案路徑>")
        return

    input_file = sys.argv[1]
    config = load_config()
    azure_ai = AzureOpenAIAPI()
    gemini = GeminiAPI()
    
    # 1. 提取 BOM
    print_progress(f"開始處理: {input_file}")
    from extract_bom_data import extract_bom_text_from_excel, extract_bom_text_from_pdf
    
    file_ext = os.path.splitext(input_file)[1].lower()
    if file_ext in ['.xlsx', '.xls']:
        data = extract_bom_text_from_excel(input_file)
    elif file_ext == '.pdf':
        data = extract_bom_text_from_pdf(input_file)
    else:
        print("❌ 不支援的檔案格式")
        return

    if not data:
        print("❌ 提取 BOM 文字失敗，請確認檔案內容。")
        return

    print(f"[INFO] 讀取到 {len(data['full_text'])} 字元的 BOM 內容")

    # 2. 生成向量並搜尋相似模板
    print_progress("生成向量...")
    vector = azure_ai.get_embedding(data['full_text'])
    if not vector:
        print("❌ 向量生成失敗。")
        return
    
    print_progress("搜尋相似模板...")
    similar = query_similar_boms(vector, config)
    if not similar:
        print("❌ 找不到相似模板。")
        return
    
    ref = similar[0]
    print(f"✅ 選用模板: {ref['filename']} (相似度: {ref['score']:.4f})")

    # 3. 生成
    print_progress("正在生成 SOP 內容...")
    steps = gemini.generate_assembly_steps(data, ref)
    if not steps:
        print("❌ AI 生成內容失敗或格式錯誤。")
        return

    # 4. 存檔
    output_dir = config['PATHS']['output_folder']
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"SOP_{os.path.basename(input_file)}")
    
    create_styled_excel(steps, os.path.basename(input_file), output_path)
    print(f"🎉 成功！檔案已儲存: {output_path}")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"❌ 執行錯誤: {e}")